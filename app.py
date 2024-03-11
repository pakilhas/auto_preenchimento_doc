from openpyxl import load_workbook
from docx import Document
from datetime import datetime

#ler os arquivos da planilha com os dados dos fornecedores
#ler a pagina da planilha (no caso Sheet1)
#passar para word as informações da planilha
planilha_forn = load_workbook('./fornecedores.xlsx')
pag_fornecedores = planilha_forn['Sheet1']

#passando por cada linha da planilha começanda(min_row = 2) pela 2 linha da planilha e cada vez que ler apenas os dados irão retornar (values_only = true)
for linha in pag_fornecedores.iter_rows(min_row =2, values_only= True):
    #passando cada informação da linha para uma variavel = unpacking
    nome_empresa, endereco, cidade, estado, cep, telefone, email, setor = linha
   
   #adicionar infomaçoes no arquivo word(doc)
    arquivo_word = Document ()
    arquivo_word.add_heading('Contrato de prestação de serviços', 0)

    #por o texto e subistituir com CHAVES o local que deseja por as informações retiradas da planilha
    texto_contrato = f"""
    Este contrato de prestação de serviços é feito entre {nome_empresa}, com endereço em {endereco}, {cidade}, {estado}, CEP {cep}, doravante denominado FORNECEDOR, e a empresa CONTRATANTE.

    Pelo presente instrumento particular, as partes têm, entre si, justo e acordado o seguinte:

    1. OBJETO DO CONTRATO
    O FORNECEDOR compromete-se a fornecer à CONTRATANTE os serviços/material de acordo com as especificações acordadas, respeitando os padrões de qualidade e os prazos estipulados.

    2. PRAZO
    Este contrato tem prazo de vigência de 12 (doze) meses, iniciando-se na data de sua assinatura, podendo ser renovado conforme acordo entre as partes.

    3. VALOR E FORMA DE PAGAMENTO
    O valor dos serviços prestados será acordado conforme as demandas da CONTRATANTE e a capacidade de entrega do FORNECEDOR. Os pagamentos serão realizados mensalmente, mediante apresentação de nota fiscal.

    4. CONFIDENCIALIDADE
    Todas as informações trocadas entre as partes durante a vigência deste contrato serão tratadas como confidenciais.

    Para firmeza e como prova de assim haverem justo e contratado, as partes assinam o presente contrato em duas vias de igual teor e forma.

    FORNECEDOR: {nome_empresa}
    E-mail: {email}

    CONTRATANTE: kiradopy
    E-mail: kiradopai@gkira.com

    [Itajai, SC],{datetime.now().strftime('%d/%m/%Y')}

    """
    #adicionar texto gerado ao word(docx)
    arquivo_word.add_paragraph(texto_contrato)
    #salvar arquivos .docx
    arquivo_word.save(f'./contratos/contrato_{nome_empresa}.docx')